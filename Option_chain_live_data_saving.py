import Intraday_live_data
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import schedule
import time 
from datetime import  timedelta, datetime
from openpyxl import Workbook
import requests

i=0
ii=0
MAX_ATTEMPTS = 50
DELAY_BETWEEN_ATTEMPTS = 5
path="D:/ashu/Finance/algo_trading/Option_chain_data/"
script="NIFTY"
Expiry_Date="31-Aug-2023"


def Time_recording():
    current_time = time.strftime("%H:%M:%S")
    # Convert the 24-hour format to 12-hour format with AM/PM
    time_parts = current_time.split(":")
    hours = int(time_parts[0])
    minutes = int(time_parts[1])
    seconds = int(time_parts[2])
    if hours >= 12:
        suffix = "PM"
        hours -= 12
    else:
        suffix = "AM"
    if hours == 0:
        hours = 12
    formatted_time = "{:02d}:{:02d}:{:02d} {}".format(hours, minutes, seconds, suffix)

    return formatted_time


def loading_data(book,ws):
    print(9)
    global ii
    global path
    global script
    global Expiry_Date

    #######################stocks and the index computation###############################################

    # stocks_and_index=["NIFTY","BANKNIFTY"]


    ################## Time Computation ########################
    formatted_time=Time_recording()
    print(formatted_time)

    # nifty_closing_price=Intraday_live_data.nifty_live_data()

    new_data = {'Name': ["OI_Data",""], 'Time': [f"{formatted_time}",""], 'Time_interval': ["5 Minutes",""]}
    df = pd.DataFrame(new_data)

    Option_chain=Intraday_live_data.getoptionchain(script,Expiry_Date)
    Option_chain=Option_chain.fillna(0)
    df_1=Option_chain
            # Write the new data to the sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for r in dataframe_to_rows(df_1, index=False, header=True):
        ws.append(r)

            # Save the Excel file
    book.save(path+f'data{ii}.xlsx')


def calling():

    global i
    global ii
    global path
    print(6)
    global script
    global Expiry_Date

    if i<1:

        i=i+1
        book = load_workbook(path+f'data{ii}.xlsx')
        print(7)
        print(f"the value of the ii is {ii}")

        ws = book[script]
        print(10)
        loading_data(book,ws)

    else:
        print(8)
        book = load_workbook(path+f'data{ii}.xlsx')
        ws =  book[script]
        loading_data(book,ws)





def running():
    print(5)
    schedule.every(5).minutes.until(timedelta(hours=6)).do(calling)

    while True:
        schedule.run_pending()
        time.sleep(1)



def make_request_again():
    global i
    global ii
    global path
    global script
    global Expiry_Date

    attempt = 1
    while attempt <= MAX_ATTEMPTS:
        try:
            print(3)
            ii=ii+1
            workbook = Workbook()
            workbook.remove(workbook.active)
            sheet1 = workbook.create_sheet("NIFTY")
            workbook.save(path+f"data{ii}.xlsx")
            book = load_workbook(path+f'data{ii}.xlsx')
            ws = book['NIFTY']

            formatted_time = Time_recording()
            print(formatted_time)

            new_data = {'Name': ["OI_Data",""], 'Time': [f"{formatted_time}",""], 'Time_interval': ["5 Minutes",""]}
            df = pd.DataFrame(new_data)

            Option_chain=Intraday_live_data.getoptionchain(script,Expiry_Date)
            Option_chain=Option_chain.fillna(0)
            df_1=Option_chain

            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            for r in dataframe_to_rows(df_1, index=False, header=True):
                ws.append(r)

                # Save the Excel file
            book.save(path+f'data{ii}.xlsx')
            running()
        except (requests.exceptions.RequestException, requests.exceptions.HTTPError) as e:
            print(4)
            i=0
            print(f"Attempt {attempt} failed: {e}")
            attempt += 1
            time.sleep(DELAY_BETWEEN_ATTEMPTS)
    
    # If the maximum number of attempts is reached without a successful response, raise an exception
    raise Exception("Maximum number of attempts reached without a successful response")





try:
    print(1)
    make_request_again()
    
except Exception as e:
    # Handle the exception here or log the error message
    print(2)
    print("Check the Expiry Date")
    print(f"Error occurred: {e}")
    